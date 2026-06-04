"""
preprocess_atomic.py

Standalone feature-engineering script for transition-metal atomic energy levels.

It reads a *raw* Kurucz/Cowan level file (XLSX or CSV) — e.g.
``data/kurucz/Co_kurucz_raw.xlsx`` — and writes a *rich* feature workbook
(``data/Co_features_rich.xlsx``) that contains EVERY computed feature as an
explicit column.  The goal is to move all on-the-fly feature engineering out of
``AtomicDataset.py`` and into a single, inspectable, version-controllable file.

Raw input columns (one row per energy level)::

    J, parity_flag, Configuration_raw, Term_raw, OBS.LEVEL, EIGENVALUE,
    T-W, calc.gJ, obs.gJ, Reference

  * ``EIGENVALUE`` is the theoretically calculated energy (Cowan code), cm⁻¹.
  * ``OBS.LEVEL``  is the experimentally observed energy (the ML TARGET), cm⁻¹.
  * ``T-W``        is the residual OBS.LEVEL − EIGENVALUE, cm⁻¹.

Usage::

    python preprocess_atomic.py --config config_atomic.yaml

------------------------------------------------------------------------------
NOTE FOR A FUTURE STEP (AtomicDataset.py refactor)
------------------------------------------------------------------------------
This script intentionally does NOT modify ``AtomicDataset.py``.  Once this
rich feature file exists, ``AtomicDataset.py`` should be simplified so that it
merely:

    1. loads ``Co_features_rich.xlsx`` (the "features" sheet),
    2. selects the configured input-feature columns *by name*,
    3. performs the train/val/test split and StandardScaler scaling.

All the parsing / physics computation currently living in
``_add_derived_features``, ``_add_transition_metal_features``,
``_add_theoretical_lande_g`` and ``_encode_valence_electrons`` is reproduced
here and baked into the output columns, so the Dataset class will no longer
need to recompute anything on instantiation.

IMPORTANT BEHAVIOUR CHANGE vs. the current AtomicDataset.py:
    The valence-slot ordering is REVERSED.  Here ``val_e1`` is the OUTERMOST
    electron (highest Madelung order); in the old code ``val_e1`` was the most
    core-like valence electron.  This fixes the naming confusion flagged by the
    permutation-importance analysis.

Author: Aga (generated with Claude Code)
"""

import argparse
import os
import re
import subprocess
import sys

import numpy as np
import pandas as pd
import yaml
from tqdm import tqdm

# Ensure openpyxl is available (used for both reading .xlsx and writing the
# richly-formatted output workbook).  Installed on demand so the script is
# self-contained on a fresh environment.
try:
    import openpyxl  # noqa: F401
except ImportError:  # pragma: no cover - environment bootstrap
    print("openpyxl not found - installing it now...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl  # noqa: F401

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ===========================================================================
# Physical / parsing constants
# ===========================================================================

# Maps a term-symbol orbital letter to its total orbital angular momentum L.
# (Russell-Saunders term notation, e.g. the 'F' in '4F' means L = 3.)
L_MAP = {'S': 0, 'P': 1, 'D': 2, 'F': 3, 'G': 4, 'H': 5, 'I': 6, 'K': 7}

# Maps a sub-shell letter to its single-electron angular momentum ℓ.
# Used for orbital occupancy parsing and parity computation.
L_VALUES = {'s': 0, 'p': 1, 'd': 2, 'f': 3}

# Full, fixed list of orbital-occupancy columns the output always contains.
ORBITAL_COLUMNS = ['1s', '2s', '2p', '3s', '3p', '3d', '4s', '4p',
                   '4d', '4f', '5s', '5p', '5d', '6s', '6p']

# Cobalt core is closed and identical for every level: [Ar] = 1s²2s²2p⁶3s²3p⁶.
# Only the valence orbitals (3d and beyond) vary between configurations.
CORE_OCCUPANCY = {'1s': 2, '2s': 2, '2p': 6, '3s': 2, '3p': 6}
CORE_COLUMNS = ['1s', '2s', '2p', '3s', '3p']  # the 18 core electrons

# Valence orbitals = everything that is not core.  Used to build valence slots.
VALENCE_COLUMNS = [o for o in ORBITAL_COLUMNS if o not in CORE_OCCUPANCY]

# Regex for a bracketed coupling term such as '(4F)' or '(2P*)'.
_BRACKET_RE = re.compile(r'\((\d[A-Z]\*?)\)')
# Regex for the comma-separated sub-resultant term such as ',(3P)'.
_SUBRES_RE = re.compile(r',\((\d[A-Z]\*?)\)')
# Regex for an orbital occupancy segment such as '3d7', '4s', '5s2'.
# Group 3 (the occupancy) is optional; absent means a single electron.
_ORBITAL_RE = re.compile(r'(\d)([spdf])(\d+)?')


# ===========================================================================
# Config loading
# ===========================================================================

def load_preprocessing_config(config_path):
    """
    Load the ``preprocessing`` section from the YAML config file.

    Returns a plain dict with sensible defaults filled in so the script runs
    even if some keys are missing.  We read YAML directly (rather than through
    utils.load_config / OmegaConf) to keep this script dependency-light.
    """
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Config file not found: {config_path}")

    with open(config_path, 'r', encoding='utf-8') as fh:
        full_cfg = yaml.safe_load(fh)

    pp = (full_cfg or {}).get('preprocessing', {}) or {}

    # Defaults match the Co I dataset described in the task.
    defaults = {
        'input_file': 'data/kurucz/Co_kurucz_raw.xlsx',
        'output_file': 'data/Co_features_rich.xlsx',
        'element': 'Co',
        'Z': 27,
        'A': 59,
        'ionization_energy': 63564.0,
        'max_valence': 9,
        'zeta_3d': 515.0,
        'inverse_target_scale': 100000.0,  # A in Inverse_Binding_Energy = A / binding
    }
    for key, val in defaults.items():
        pp.setdefault(key, val)

    return pp


# ===========================================================================
# TASK 1 - configuration-string parsing into component term features
# ===========================================================================

def parse_term_SL(term):
    """
    Parse a single Russell-Saunders term string into (S, L).

    Physics: a term symbol ``(2S+1)L`` encodes the total spin S via its
    multiplicity (2S+1) and the total orbital momentum L via its letter.
    Example: ``'4F'`` → multiplicity 4 → S = (4-1)/2 = 1.5; letter 'F' → L = 3.

    A trailing ``'*'`` marks odd parity and is stripped before parsing.

    Returns:
        (S, L) as floats, or (np.nan, np.nan) if the string cannot be parsed.
    """
    if term is None:
        return np.nan, np.nan
    t = str(term).strip()
    if t.endswith('*'):            # strip the odd-parity marker
        t = t[:-1]
    m = re.match(r'(\d)([A-Z])', t)
    if not m:
        return np.nan, np.nan
    multiplicity = int(m.group(1))         # 2S + 1
    S = (multiplicity - 1) / 2.0           # total spin
    L = L_MAP.get(m.group(2), np.nan)      # total orbital angular momentum
    return float(S), float(L)


def parse_configuration(config_str, term_str):
    """
    Parse a configuration string and resultant term into coupling features.

    Configuration strings such as ``'3d7(4F)4s(2S)4p(2P),(3P)'`` describe how
    sub-groups of electrons couple their angular momenta.  The bracketed terms
    appear in order:

        (3d-group term)  (outer-electron-1 term)  (outer-electron-2 term)
        and optionally, after a comma, the ,(sub-resultant) term that couples
        the last two outer electrons before they combine with the 3d group.

    The comma sub-resultant is matched a second time by the generic bracket
    regex, so it is removed from the tail of the main-bracket list.

    Each component is mapped to a fixed-width set of 8 slots (padding with 0.0):

        comp1_S/comp1_L   3d-group coupling      (present except for '3d9')
        comp2_S/comp2_L   first outer electron   (4s, 4p, ...)
        comp3_S/comp3_L   second outer electron  (3-bracket configs only)
        subres_S/subres_L sub-resultant ,(xY)

    The resultant term (``Term_raw``) is parsed separately into
    result_S / result_L — the final coupled state of the whole atom.

    Returns a dict with the 8 component slots, result_S/result_L,
    n_components and has_subresultant.
    """
    cfg = '' if config_str is None else str(config_str)

    # Step 1+2: pull out the sub-resultant first, then all brackets, and drop
    # the sub-resultant's duplicate occurrence from the main list.
    sub_terms = _SUBRES_RE.findall(cfg)            # e.g. ['3P'] or []
    all_brackets = _BRACKET_RE.findall(cfg)        # includes the sub-resultant
    has_sub = 1 if sub_terms else 0
    if has_sub:
        # The sub-resultant is the final bracket captured by the generic regex.
        main_brackets = all_brackets[:-1]
        subres_term = sub_terms[0]
    else:
        main_brackets = all_brackets
        subres_term = None

    n_components = len(main_brackets)              # 0, 1, 2, or 3

    # Step 3+4: parse each main bracket into (S, L) and place into fixed slots.
    # Default every slot to 0.0 so the output width is constant.
    slots = {
        'comp1_S': 0.0, 'comp1_L': 0.0,
        'comp2_S': 0.0, 'comp2_L': 0.0,
        'comp3_S': 0.0, 'comp3_L': 0.0,
        'subres_S': 0.0, 'subres_L': 0.0,
    }
    for i, bracket in enumerate(main_brackets[:3]):   # at most 3 main components
        S, L = parse_term_SL(bracket)
        # NaN (unparseable) collapses to the 0.0 padding convention.
        slots[f'comp{i + 1}_S'] = 0.0 if pd.isna(S) else S
        slots[f'comp{i + 1}_L'] = 0.0 if pd.isna(L) else L

    if has_sub:
        S, L = parse_term_SL(subres_term)
        slots['subres_S'] = 0.0 if pd.isna(S) else S
        slots['subres_L'] = 0.0 if pd.isna(L) else L

    # Step 5: parse the resultant (overall) term symbol.
    result_S, result_L = parse_term_SL(term_str)

    slots['result_S'] = result_S
    slots['result_L'] = result_L
    slots['n_components'] = int(n_components)
    slots['has_subresultant'] = int(has_sub)
    return slots


# ===========================================================================
# TASK 2 - orbital electron counts
# ===========================================================================

def parse_orbitals(config_str):
    """
    Parse per-orbital electron counts from a configuration string.

    The bracketed coupling terms use UPPERCASE letters (e.g. '(4F)'), while
    orbital sub-shells use lowercase ([spdf]); the regex therefore matches only
    genuine orbital segments and skips the coupling brackets automatically.
    An orbital written without an explicit count (e.g. '4s') holds 1 electron.

    The cobalt core (1s²2s²2p⁶3s²3p⁶ = 18 e⁻) is constant and is filled in
    directly; only the valence orbitals come from the string.

    Returns a dict mapping every column in ORBITAL_COLUMNS to its occupancy.
    """
    occ = {o: 0 for o in ORBITAL_COLUMNS}
    occ.update(CORE_OCCUPANCY)                 # closed [Ar] core, always present

    cfg = '' if config_str is None else str(config_str)
    for m in _ORBITAL_RE.finditer(cfg):
        n, letter, count = m.group(1), m.group(2), m.group(3)
        key = f'{n}{letter}'                   # e.g. '3d'
        occupancy = int(count) if count else 1  # bare orbital → 1 electron
        if key in occ:
            occ[key] = occupancy
    return occ


# ===========================================================================
# TASK 3 - physics-derived features
# ===========================================================================

def compute_zeff_3d(occ, Z):
    """
    Effective nuclear charge seen by a 3d electron via Slater's rules.

    σ = 0.35·(n_3d − 1)            other electrons in the same 3d group
      + 0.85·(n_3s + n_3p)         the n−1 shell (moderate screening)
      + 1.00·(n_1s + n_2s + n_2p)  deeper shells (full screening)
    Z_eff = Z − σ.  Electrons in shells outside 3d (4s, 4p, …) do not screen it.

    Z_eff² is the dominant energy scale, analogous to Z²/n² in hydrogen.
    """
    n_3d = occ['3d']
    sigma = 0.35 * max(n_3d - 1, 0)               # same-group screening
    sigma += 0.85 * (occ['3s'] + occ['3p'])       # n-1 shell
    sigma += 1.00 * (occ['1s'] + occ['2s'] + occ['2p'])  # inner shells
    return Z - sigma


def compute_valence_slots(occ, max_valence):
    """
    Build the fixed-width valence-slot feature vector (outermost-first).

    Every valence electron is expanded to an individual (n, ℓ) pair, then
    sorted by Madelung filling order — key (n+ℓ, n), where a lower value fills
    first — and the order is REVERSED so slot 1 holds the outermost electron
    (highest Madelung order).  The list is padded with (0, 0) up to max_valence.

    This outermost-first convention is the deliberate fix to the slot-ordering
    confusion identified in the permutation-importance analysis (the old code
    placed the outermost electron in the LAST slot).

    Returns a flat list [val_e1_n, val_e1_l, ..., val_e{max}_n, val_e{max}_l].
    """
    electrons = []
    for o in VALENCE_COLUMNS:
        n = int(o[0])
        l = L_VALUES[o[1]]
        electrons.extend([(n, l)] * int(occ[o]))   # one (n, ℓ) per electron

    # Madelung order ascending (fills first); then reverse → outermost first.
    electrons.sort(key=lambda nl: (nl[0] + nl[1], nl[0]))
    electrons.reverse()

    slots = electrons[:max_valence]                 # never overflow the slots
    while len(slots) < max_valence:                 # right-pad with (0, 0)
        slots.append((0, 0))

    flat = []
    for n, l in slots:
        flat.extend([float(n), float(l)])
    return flat


def compute_row_features(row, cfg):
    """
    Compute the full feature dict for one raw input row.

    Combines TASK 1 (component terms), TASK 2 (orbital occupancies) and
    TASK 3 (physics-derived features) plus TASK 4 (gJ handling) and TASK 5
    (atomic constants) into a single flat dict, one entry per output column.
    """
    Z = cfg['Z']
    A = cfg['A']
    max_valence = cfg['max_valence']
    zeta = cfg['zeta_3d']

    feat = {}

    # --- Identifiers (passed straight through) ---------------------------
    feat['Configuration_raw'] = row.get('Configuration_raw')
    feat['Term_raw'] = row.get('Term_raw')
    feat['J'] = float(row['J'])
    feat['parity_flag'] = int(row['parity_flag'])
    feat['Reference'] = row.get('Reference')

    # --- Raw energies (TASK 3i) -----------------------------------------
    obs_level = float(row['OBS.LEVEL'])
    eigenvalue = float(row['EIGENVALUE'])
    feat['OBS.LEVEL'] = obs_level                  # experimental TARGET
    feat['EIGENVALUE'] = eigenvalue                # raw Cowan baseline (eigenvalue_calc)
    # delta_e_theory: prefer the tabulated T-W residual; otherwise derive it.
    tw = row.get('T-W', np.nan)
    feat['delta_e_theory'] = float(tw) if pd.notna(tw) else (obs_level - eigenvalue)

    # --- gJ handling (TASK 4) -------------------------------------------
    obs_gj = row.get('obs.gJ', np.nan)
    calc_gj = row.get('calc.gJ', np.nan)
    obs_gj = float(obs_gj) if pd.notna(obs_gj) else np.nan   # NaN kept as NaN!
    calc_gj = float(calc_gj) if pd.notna(calc_gj) else np.nan
    feat['obs_gj'] = obs_gj                                  # multi-task TARGET
    feat['has_obs_gj'] = int(pd.notna(obs_gj))              # mask for the loss
    feat['calc_gj'] = calc_gj                                # INPUT feature
    feat['obs_minus_calc_gj'] = (obs_gj - calc_gj) if pd.notna(obs_gj) else np.nan

    # --- TASK 1: component term features --------------------------------
    comp = parse_configuration(row.get('Configuration_raw'), row.get('Term_raw'))
    result_S = comp['result_S']
    result_L = comp['result_L']
    feat['result_S'] = result_S
    feat['result_L'] = result_L
    feat['term_known'] = int(pd.notna(result_S) and pd.notna(result_L))

    # --- TASK 3a: angular-momentum products -----------------------------
    J = feat['J']
    J_sq = J * (J + 1)
    L_sq = result_L * (result_L + 1) if pd.notna(result_L) else np.nan
    S_sq = result_S * (result_S + 1) if pd.notna(result_S) else np.nan
    lande_so_term = J_sq - L_sq - S_sq             # ∝ spin-orbit shift
    feat['J_sq'] = J_sq
    feat['L_sq'] = L_sq if pd.notna(L_sq) else 0.0
    feat['S_sq'] = S_sq if pd.notna(S_sq) else 0.0
    feat['lande_so_term'] = lande_so_term if pd.notna(lande_so_term) else 0.0

    # --- TASK 3b: theoretical Landé g-factor (LS coupling) --------------
    # WARNING: lande_g_theoretical uses the pure LS-coupling Landé formula.
    # For alkali atoms (single valence electron, pure LS coupling) this is accurate.
    # For iron-group atoms (Co, Fe, Ni) with strong configuration mixing and
    # intermediate coupling, this formula gives values far from experiment.
    # DO NOT use as a model input feature for transition metal atoms.
    # Magda's calc_gJ (from Cowan code diagonalization) is the correct theoretical value,
    # but it is semi-empirical and must not be used as a training feature either
    if J == 0:
        lande_g_theo = 0.0                          # formula undefined; g_J ≡ 0
        has_lande = 1 if feat['term_known'] else 0
    elif pd.isna(result_S) or pd.isna(result_L):
        lande_g_theo = np.nan                        # cannot evaluate
        has_lande = 0
    else:
        lande_g_theo = 1.0 + lande_so_term / (2.0 * J_sq)
        has_lande = 1
    feat['has_lande_theoretical'] = has_lande
    feat['lande_g_theoretical'] = 0.0 if pd.isna(lande_g_theo) else lande_g_theo

    # --- TASK 1 (cont.): component slots --------------------------------
    for k in ['comp1_S', 'comp1_L', 'comp2_S', 'comp2_L', 'comp3_S', 'comp3_L',
              'subres_S', 'subres_L']:
        feat[k] = comp[k]
    feat['n_components'] = comp['n_components']
    feat['has_subresultant'] = comp['has_subresultant']

    # --- TASK 2: orbital occupancies ------------------------------------
    occ = parse_orbitals(row.get('Configuration_raw'))

    # --- TASK 3c: d-electron features -----------------------------------
    n_3d = occ['3d']
    feat['n_3d'] = float(n_3d)
    feat['d_holes'] = float(10 - n_3d)             # electron-hole symmetry
    feat['d_from_half'] = float(abs(n_3d - 5))     # exchange-energy pairing cost
    feat['is_half_filled'] = int(n_3d == 5)        # special Hund stability

    # --- TASK 3d: effective nuclear charge ------------------------------
    z_eff = compute_zeff_3d(occ, Z)
    feat['Z_eff'] = z_eff
    feat['Z_eff_sq'] = z_eff ** 2

    # --- TASK 3e: spin-orbit energy estimate ----------------------------
    feat['zeta_3d'] = float(zeta)
    feat['E_so_estimate'] = (zeta / 2.0) * feat['lande_so_term']

    # --- TASK 3f: parity computed from orbital occupancy ----------------
    parity = sum(occ[o] * L_VALUES[o[1]] for o in ORBITAL_COLUMNS) % 2
    feat['parity_computed'] = int(parity)          # 0 = even, 1 = odd

    # --- TASK 3h: valence slots (outermost-first) -----------------------
    slot_vals = compute_valence_slots(occ, max_valence)
    for i in range(max_valence):
        feat[f'val_e{i + 1}_n'] = slot_vals[2 * i]
        feat[f'val_e{i + 1}_l'] = slot_vals[2 * i + 1]

    # --- TASK 3g: valence summary ---------------------------------------
    total_e = sum(occ[o] for o in ORBITAL_COLUMNS)
    core_e = sum(occ[o] for o in CORE_COLUMNS)
    feat['valence_electrons'] = float(total_e - core_e)
    feat['total_electrons'] = float(total_e)
    feat['core_electrons'] = float(core_e)
    occupied_n = [int(o[0]) for o in ORBITAL_COLUMNS if occ[o] > 0]
    feat['max_principal_n'] = float(max(occupied_n)) if occupied_n else 0.0

    # --- TASK 2 (cont.): raw orbital occupancy columns ------------------
    for o in ORBITAL_COLUMNS:
        feat[o] = float(occ[o])

    # --- TASK 5: atomic constants ---------------------------------------
    feat['Z'] = int(Z)
    feat['A'] = int(A)
    feat['proton_number'] = int(Z)
    feat['neutron_number'] = int(A - Z)

    return feat


# ===========================================================================
# Output column ordering (TASK 6)
# ===========================================================================

def add_element_targets_and_rydberg(df, cfg):
    """
    Add the element label, pre-computed energy-target transforms, and Rydberg
    placeholder columns to the features DataFrame (in place).

    These columns let AtomicDataset.py select a target purely by name and toggle
    feature groups without recomputing anything:

      Element                      element symbol (needed for stratified splits
                                   and multi-element merging downstream).
      Binding_Energy_cm-1          E_ion - OBS.LEVEL, clipped at 0 (continuum).
      Log_Binding_Energy_cm-1      log(binding); NaN where binding <= 0.
      Inverse_Binding_Energy_cm-1  A / binding; NaN where binding <= 0.
      n_star, rydberg_prediction, one_over_nstar_sq
                                   Rydberg series features. These are meaningful
                                   only for single-valence-electron (alkali) atoms;
                                   for transition metals such as Co they are 0.0.
    """
    E_ion = float(cfg['ionization_energy'])          # ionization energy (cm^-1)
    scale = float(cfg['inverse_target_scale'])       # A in A / binding

    df['Element'] = cfg['element']                    # constant for a single-element file

    # Binding energy = how far below the ionization limit the level sits.
    binding = (E_ion - df['OBS.LEVEL']).clip(lower=0.0)
    df['Binding_Energy_cm-1'] = binding

    # Log / inverse transforms are undefined at binding == 0 (continuum) → NaN.
    positive = binding.where(binding > 0)
    with np.errstate(divide='ignore', invalid='ignore'):
        df['Log_Binding_Energy_cm-1'] = np.log(positive)
        df['Inverse_Binding_Energy_cm-1'] = scale / positive

    # Rydberg features are not applicable to transition metals → zeros.
    df['n_star'] = 0.0
    df['rydberg_prediction'] = 0.0
    df['one_over_nstar_sq'] = 0.0


def build_column_order(max_valence):
    """Return the explicit output column order described in TASK 6."""
    valence_cols = []
    for i in range(max_valence):
        valence_cols.extend([f'val_e{i + 1}_n', f'val_e{i + 1}_l'])

    return (
        # Identifiers (not model inputs)
        ['Configuration_raw', 'Term_raw', 'J', 'parity_flag', 'Reference', 'Element']
        # Raw energy (target + Cowan baseline) + pre-computed target transforms
        + ['OBS.LEVEL', 'EIGENVALUE', 'delta_e_theory',
           'Binding_Energy_cm-1', 'Log_Binding_Energy_cm-1', 'Inverse_Binding_Energy_cm-1']
        # Experimental gJ (multi-task target)
        + ['obs_gj', 'has_obs_gj']
        # Theoretical gJ (input feature)
        + ['calc_gj', 'lande_g_theoretical', 'has_lande_theoretical', 'obs_minus_calc_gj']
        # Parsed quantum numbers (resultant term)
        + ['result_S', 'result_L', 'term_known']
        # Angular momentum products
        + ['J_sq', 'L_sq', 'S_sq', 'lande_so_term']
        # Component term features (new)
        + ['comp1_S', 'comp1_L', 'comp2_S', 'comp2_L', 'comp3_S', 'comp3_L',
           'subres_S', 'subres_L', 'n_components', 'has_subresultant']
        # d-electron features
        + ['n_3d', 'd_holes', 'd_from_half', 'is_half_filled']
        # Screening
        + ['Z_eff', 'Z_eff_sq']
        # Spin-orbit
        + ['zeta_3d', 'E_so_estimate']
        # Parity
        + ['parity_computed']
        # Valence slots (outermost-first)
        + valence_cols
        # Valence summary
        + ['valence_electrons', 'total_electrons', 'core_electrons', 'max_principal_n']
        # Rydberg features (zeros for transition metals; populated for alkalis)
        + ['n_star', 'rydberg_prediction', 'one_over_nstar_sq']
        # Orbital occupancies (raw)
        + ORBITAL_COLUMNS
        # Atomic constants
        + ['Z', 'A', 'proton_number', 'neutron_number']
    )


# ===========================================================================
# Summary sheet (TASK 6, sheet 2)
# ===========================================================================

# Leading identifier columns are excluded from the numeric summary.
IDENTIFIER_COLUMNS = ['Configuration_raw', 'Term_raw', 'J', 'parity_flag',
                      'Reference', 'Element']


def build_summary_df(df):
    """
    Build the per-feature summary table (sheet 2).

    One row per numeric feature with: feature_name, dtype, n_non_null,
    n_unique, mean, std, min, max.  Identifier columns are skipped.
    """
    rows = []
    for col in df.columns:
        if col in IDENTIFIER_COLUMNS:
            continue
        series = df[col]
        if not pd.api.types.is_numeric_dtype(series):
            continue
        rows.append({
            'feature_name': col,
            'dtype': str(series.dtype),
            'n_non_null': int(series.notna().sum()),
            'n_unique': int(series.nunique(dropna=True)),
            'mean': series.mean(),
            'std': series.std(),
            'min': series.min(),
            'max': series.max(),
        })
    return pd.DataFrame(rows)


# ===========================================================================
# XLSX writing + formatting (TASK 6)
# ===========================================================================

# Colour palette (ARGB without the leading '#').
FILL_HEADER = PatternFill('solid', fgColor='DBEAFE')   # light blue
FILL_IDENT = PatternFill('solid', fgColor='F3F4F6')    # light gray
FILL_TARGET = PatternFill('solid', fgColor='FEF9C3')   # light yellow
FILL_COMPONENT = PatternFill('solid', fgColor='D1FAE5')  # light green
FILL_EIGEN = PatternFill('solid', fgColor='EDE9FE')    # light purple

TARGET_COLUMNS = ['OBS.LEVEL', 'obs_gj']
COMPONENT_COLUMNS = ['comp1_S', 'comp1_L', 'comp2_S', 'comp2_L', 'comp3_S',
                     'comp3_L', 'subres_S', 'subres_L', 'n_components',
                     'has_subresultant']
EIGEN_COLUMNS = ['EIGENVALUE', 'delta_e_theory']


def write_rich_xlsx(features_df, summary_df, output_path):
    """
    Write the features + summary sheets and apply all TASK 6 formatting.

    Formatting:
      * header row bold, light-blue fill, frozen;
      * identifier columns light gray, target columns yellow,
        component columns green, eigenvalue/delta columns purple;
      * numeric cells formatted to 4 decimal places;
      * approximate auto-fit column widths.
    """
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        features_df.to_excel(writer, sheet_name='features', index=False)
        summary_df.to_excel(writer, sheet_name='summary', index=False)

        wb = writer.book
        ws = wb['features']

        columns = list(features_df.columns)
        numeric_flags = [pd.api.types.is_numeric_dtype(features_df[c]) for c in columns]

        # Per-column body fill (header is overridden to blue below).
        def body_fill(col):
            if col in TARGET_COLUMNS:
                return FILL_TARGET
            if col in COMPONENT_COLUMNS:
                return FILL_COMPONENT
            if col in EIGEN_COLUMNS:
                return FILL_EIGEN
            if col in IDENTIFIER_COLUMNS:
                return FILL_IDENT
            return None

        n_rows = len(features_df)
        for c_idx, col in enumerate(columns, start=1):
            letter = get_column_letter(c_idx)
            fill = body_fill(col)
            is_numeric = numeric_flags[c_idx - 1]

            # Header cell: bold + blue, regardless of column group.
            header_cell = ws.cell(row=1, column=c_idx)
            header_cell.font = Font(bold=True)
            header_cell.fill = FILL_HEADER
            header_cell.alignment = Alignment(horizontal='center')

            # Body cells: group fill + 4-decimal number format for numerics.
            for r_idx in range(2, n_rows + 2):
                cell = ws.cell(row=r_idx, column=c_idx)
                if fill is not None:
                    cell.fill = fill
                if is_numeric:
                    cell.number_format = '0.0000'

            # Approximate auto-fit width.
            ws.column_dimensions[letter].width = max(len(str(col)), 8) * 1.2

        # Freeze the header row so it stays visible while scrolling.
        ws.freeze_panes = 'A2'

        # Light formatting for the summary sheet header too.
        ws_sum = wb['summary']
        for c_idx in range(1, len(summary_df.columns) + 1):
            cell = ws_sum.cell(row=1, column=c_idx)
            cell.font = Font(bold=True)
            cell.fill = FILL_HEADER
            ws_sum.column_dimensions[get_column_letter(c_idx)].width = 16
        ws_sum.freeze_panes = 'A2'


# ===========================================================================
# TASK 7 - validation + summary printout
# ===========================================================================

def print_validation_report(df, cfg):
    """Print the validation / coverage report described in TASK 7."""
    n = len(df)
    print("\n" + "=" * 60)
    print("VALIDATION REPORT")
    print("=" * 60)
    print(f"Total rows: {n}")

    n_obs_gj = int(df['has_obs_gj'].sum())
    print(f"Rows with obs_gj: {n_obs_gj} / {n} ({100.0 * n_obs_gj / n:.1f}%)")

    n_term = int(df['term_known'].sum())
    print(f"Rows with term_known: {n_term} / {n}")

    # Electron-count check (must equal Z = 27 for every Co row).
    expected_e = cfg['Z']
    bad_counts = df[df['total_electrons'] != expected_e]
    all_ok = len(bad_counts) == 0
    print(f"Electron count check: all rows = {expected_e}? "
          f"{'yes' if all_ok else 'no'}")
    if not all_ok:
        for _, r in bad_counts.iterrows():
            print(f"  WARNING: total_electrons={r['total_electrons']:.0f} "
                  f"for config '{r['Configuration_raw']}'")

    # Parity consistency (computed vs the raw parity_flag).
    parity_match = int((df['parity_computed'] == df['parity_flag']).sum())
    print(f"Parity consistency (parity_computed vs parity_flag): "
          f"{parity_match} / {n} match")

    # Component-term coverage.
    print("Component term coverage:")
    counts = df['n_components'].value_counts().to_dict()
    print(f"  0-bracket configs: {int(counts.get(0, 0))}")
    print(f"  1-bracket configs: {int(counts.get(1, 0))}")
    print(f"  2-bracket configs: {int(counts.get(2, 0))}")
    n_three_sub = int(((df['n_components'] == 3) & (df['has_subresultant'] == 1)).sum())
    print(f"  3-bracket + sub-resultant: {n_three_sub}")

    # Top 5 most common 3d-group terms (comp1 as a multiplicity/L pair).
    print("\nTop 5 most common 3d-group terms (comp1):")
    inv_L = {v: k for k, v in L_MAP.items()}

    def comp1_label(r):
        # Reconstruct the term symbol from comp1_S / comp1_L for display.
        S, L = r['comp1_S'], r['comp1_L']
        if S == 0 and L == 0 and r['n_components'] == 0:
            return '(none: 3d9)'
        mult = int(round(2 * S + 1))
        return f"{mult}{inv_L.get(int(L), '?')}"

    labels = df.apply(comp1_label, axis=1)
    for term, cnt in labels.value_counts().head(5).items():
        print(f"  {term}: {cnt} rows")


# ===========================================================================
# Main pipeline
# ===========================================================================

def read_raw_input(input_file):
    """Read the raw input file, dispatching on the .xlsx/.csv extension."""
    ext = os.path.splitext(input_file)[1].lower()
    if ext in ('.xlsx', '.xlsm', '.xls'):
        return pd.read_excel(input_file)
    elif ext == '.csv':
        return pd.read_csv(input_file)
    else:
        raise ValueError(f"Unsupported input extension '{ext}' (expected .xlsx or .csv)")


def main():
    parser = argparse.ArgumentParser(
        description="Build a rich atomic-feature XLSX from a raw level file."
    )
    parser.add_argument('--config', type=str, default='config_atomic.yaml',
                        help="Path to the YAML config (default: config_atomic.yaml)")
    args = parser.parse_args()

    cfg = load_preprocessing_config(args.config)

    input_file = cfg['input_file']
    output_file = cfg['output_file']
    max_valence = cfg['max_valence']

    print("=" * 60)
    print("ATOMIC FEATURE PREPROCESSING")
    print("=" * 60)
    print(f"Element        : {cfg['element']}  (Z={cfg['Z']}, A={cfg['A']})")
    print(f"Input file     : {input_file}")
    print(f"Output file    : {output_file}")
    print(f"max_valence    : {max_valence}")
    print(f"zeta_3d        : {cfg['zeta_3d']} cm^-1")

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Input file not found: {input_file}")

    raw_df = read_raw_input(input_file)
    print(f"\nLoaded {len(raw_df)} raw rows with columns: {list(raw_df.columns)}")

    # Row-by-row feature computation with a progress bar.
    records = []
    for _, row in tqdm(raw_df.iterrows(), total=len(raw_df), desc="Computing features"):
        records.append(compute_row_features(row, cfg))

    features_df = pd.DataFrame.from_records(records)

    # Add element label + pre-computed target transforms + Rydberg placeholders.
    add_element_targets_and_rydberg(features_df, cfg)

    # Enforce the TASK 6 column order (any unexpected columns appended at end).
    column_order = build_column_order(max_valence)
    ordered = [c for c in column_order if c in features_df.columns]
    extra = [c for c in features_df.columns if c not in column_order]
    features_df = features_df[ordered + extra]

    # Component-term distribution summary (TASK 1, step 4).
    print("\nComponent distribution:")
    print(f"  n_components   : {features_df['n_components'].value_counts().sort_index().to_dict()}")
    print(f"  has_subresultant: {features_df['has_subresultant'].value_counts().sort_index().to_dict()}")

    summary_df = build_summary_df(features_df)

    write_rich_xlsx(features_df, summary_df, output_file)

    print_validation_report(features_df, cfg)

    print(f"\nOutput written to: {output_file} "
          f"({len(features_df)} rows, {len(features_df.columns)} columns)")


if __name__ == '__main__':
    main()
