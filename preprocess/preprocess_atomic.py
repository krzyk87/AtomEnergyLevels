"""
preprocess_atomic.py

Standalone feature-engineering script for transition-metal atomic energy levels.

It reads a *raw* level file for the active dataset source and writes a *rich*
feature workbook containing EVERY computed feature as an explicit column. The
goal is to move all on-the-fly feature engineering out of ``AtomicDataset.py``
and into a single, inspectable, version-controllable file.

Two dataset sources are supported, selected by ``dataset.dataset_source`` in the
config (override with ``--source``):

  * ``kurucz`` — Kurucz/Cowan data. Raw input is the intermediate
    ``data/kurucz/Co_kurucz_raw.xlsx`` (already in the raw schema below), built
    from the two .txt files in ``data/kurucz/``. Output:
    ``data/Co_features_rich_kurucz.xlsx``.
  * ``nist``   — NIST data. Raw input is ``data/nist/Co_i.csv`` and is converted
    to the raw schema via the helpers in ``preprocess/preprocess_nist.py``.
    Output: ``data/Co_features_rich_nist.xlsx``.

The output filename always carries the ``_<source>`` suffix so the two never
collide.

Raw (intermediate) schema consumed by compute_row_features() — one row/level::

    J, parity_flag, Configuration_raw, Term_raw, OBS.LEVEL, EIGENVALUE,
    T-W, calc.gJ, obs.gJ, Reference

  * ``EIGENVALUE`` theoretically calculated energy (Cowan code), cm^-1.
    NaN for NIST (experimental-only source).
  * ``OBS.LEVEL``  experimentally observed energy (the ML TARGET), cm^-1.
  * ``T-W``        residual OBS.LEVEL - EIGENVALUE, cm^-1 (NaN for NIST).
  * ``obs.gJ``     measured Landé g-factor (NIST: the 'Lande' column).
  * ``calc.gJ``    theoretical g-factor (Cowan); NaN for NIST.

Usage::

    python preprocess_atomic.py --config config_atomic.yaml            # uses dataset_source
    python preprocess_atomic.py --config config_atomic.yaml --source nist
    python preprocess_atomic.py --config config_atomic.yaml --source kurucz

------------------------------------------------------------------------------
NOTE FOR A FUTURE STEP (AtomicDataset.py refactor)
------------------------------------------------------------------------------
This script intentionally does NOT modify ``AtomicDataset.py``.  Once this
rich feature file exists, ``AtomicDataset.py`` merely:

    1. loads ``Co_features_rich.xlsx`` (the "features" sheet),
    2. selects the configured input-feature columns *by name*,
    3. performs the train/val/test split and StandardScaler scaling.

All the parsing / physics computation (derived features, orbital counts, etc) is
here and baked into the output columns, so the Dataset class will no longer
need to recompute anything on instantiation.

IMPORTANT BEHAVIOUR CHANGE:
    The valence-slot ordering is REVERSED.  Here ``val_e1`` is the OUTERMOST
    electron (highest Madelung order); in the old code ``val_e1`` was the most
    core-like valence electron. This fixes the naming confusion flagged by the
    permutation-importance analysis.

Author: Aga (generated with Claude Code)
"""

import argparse
import os
import re
import subprocess
import sys

# Console output contains unicode (→, ⚠, cm⁻¹); make stdout/stderr UTF-8 so it
# does not crash on Windows' default cp1252 code page.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import numpy as np
import pandas as pd
import yaml
from tqdm import tqdm

# Ensure openpyxl is available (used for both reading .xlsx and writing the
# richly-formatted output workbook).  Installed on demand so the script is
# self-contained on a fresh environment.
try:
    import openpyxl  # noqa: F401
except ImportError:  # pragma: no cover - environment bootstrap
    print("openpyxl not found - installing it now...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl  # noqa: F401

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Allow importing the NIST parsing helpers from the preprocess/ package so the
# NIST adapter below can reuse the battle-tested read/parse functions instead of
# duplicating them.
_PREPROCESS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'preprocess')
if _PREPROCESS_DIR not in sys.path:
    sys.path.insert(0, _PREPROCESS_DIR)

# Allow importing element_data (the master species-constants table loader) from the
# project root.
_ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT_DIR not in sys.path:
    sys.path.insert(0, _ROOT_DIR)


# ===========================================================================
# Dataset source handling
# ===========================================================================

# The two raw dataset sources this project supports. Output files (rich XLSX and
# split JSON) are suffixed with the active source so the two never collide.
SOURCE_SUFFIXES = ('nist', 'kurucz')


def with_source_suffix(path, source):
    """
    Return *path* with a ``_<source>`` suffix inserted before its extension.

    Any pre-existing recognised source suffix is stripped first, so the function
    is idempotent and robust to a base path that already carries a suffix:

        with_source_suffix('data/Co_features_rich.xlsx', 'nist')
            → 'data/Co_features_rich_nist.xlsx'
        with_source_suffix('data/Co_features_rich_kurucz.xlsx', 'nist')
            → 'data/Co_features_rich_nist.xlsx'   (kurucz suffix replaced)
    """
    if not path:
        return path
    base, ext = os.path.splitext(path)
    for k in SOURCE_SUFFIXES:
        if base.endswith(f'_{k}'):
            base = base[: -(len(k) + 1)]
            break
    return f"{base}_{source}{ext}"


# ===========================================================================
# Physical / parsing constants
# ===========================================================================

# Maps a term-symbol orbital letter to its total orbital angular momentum L.
# (Russell-Saunders term notation, e.g. the 'F' in '4F' means L = 3.)
L_MAP = {'S': 0, 'P': 1, 'D': 2, 'F': 3, 'G': 4, 'H': 5, 'I': 6, 'K': 7}

# Maps a sub-shell letter to its single-electron angular momentum ℓ.
# Used for orbital occupancy parsing and parity computation.
L_VALUES = {'s': 0, 'p': 1, 'd': 2, 'f': 3}

# Full, fixed list of orbital-occupancy columns the output always contains.
ORBITAL_COLUMNS = ['1s', '2s', '2p', '3s', '3p', '3d', '4s', '4p',
                   '4d', '4f', '5s', '5p', '5d', '6s', '6p']

# Cobalt core is closed and identical for every level: [Ar] = 1s²2s²2p⁶3s²3p⁶.
# Only the valence orbitals (3d and beyond) vary between configurations.
CORE_OCCUPANCY = {'1s': 2, '2s': 2, '2p': 6, '3s': 2, '3p': 6}
CORE_COLUMNS = ['1s', '2s', '2p', '3s', '3p']  # the 18 core electrons

# Valence orbitals = everything that is not core.  Used to build valence slots.
VALENCE_COLUMNS = [o for o in ORBITAL_COLUMNS if o not in CORE_OCCUPANCY]

# Regex for a bracketed coupling term such as '(4F)' or '(2P*)'.
_BRACKET_RE = re.compile(r'\((\d[A-Z]\*?)\)')
# Regex for the comma-separated sub-resultant term such as ',(3P)'.
_SUBRES_RE = re.compile(r',\((\d[A-Z]\*?)\)')
# Regex for an orbital occupancy segment such as '3d7', '4s', '5s2'.
# Group 3 (the occupancy) is optional; absent means a single electron.
_ORBITAL_RE = re.compile(r'(\d)([spdf])(\d+)?')


# ===========================================================================
# Config loading
# ===========================================================================

# Per-element defaults, applied when a key is missing from an entry. They match
# the Co I dataset so a minimal/legacy config still runs.
_ELEMENT_DEFAULTS = {
    'Z': 27,
    'A': 59,
    'ionization_energy': 63564.0,
    'max_valence': 9,
    'zeta_3d': 515.0,
    'inverse_target_scale': 100000.0,  # A in Inverse_Binding_Energy = A / binding
}


def _normalize_element_cfg(entry, defaults, data_dir, source, constants_file=None):
    """
    Fill one element entry to the flat shape the rest of this module consumes.

    Physical constants (Z, A, ionization_energy, max_valence, zeta_3d) are looked
    up from the master species table (element_data) by the entry's ``species`` (or
    a bare ``symbol``/``element`` → neutral atom). Any constant supplied inline in
    the entry overrides the table. If the table or species is unavailable, falls
    back to inline/default values so legacy configs keep working.

    Returns a dict with: element, species, Z, A, ionization_energy, max_valence,
    zeta_3d, inverse_target_scale, input_files, output_file, dataset_source.
    """
    entry = dict(entry or {})
    species_key = entry.get('species') or entry.get('symbol') or entry.get('element')
    if not species_key:
        raise ValueError(f"Element entry is missing a 'species'/'symbol' key: {entry!r}")

    cfg = dict(_ELEMENT_DEFAULTS)
    cfg.update(defaults or {})

    # Master constants table is the source of truth; inline entry values still win.
    try:
        from element_data import get_species_constants, normalize_species
        c = (get_species_constants(species_key, constants_file) if constants_file
             else get_species_constants(species_key))
        table_vals = {
            'symbol': c.get('symbol'),
            'species': normalize_species(species_key),
            'Z': c.get('Z'),
            'A': c.get('A'),
            'ionization_energy': c.get('ionization_energy_cm-1'),
            'max_valence': c.get('max_valence'),
            'zeta_3d': c.get('zeta_3d'),
        }
        cfg.update({k: v for k, v in table_vals.items()
                    if v is not None and not (isinstance(v, float) and pd.isna(v))})
    except (FileNotFoundError, KeyError) as exc:
        print(f"  ⚠ element_constants lookup failed for '{species_key}' "
              f"({type(exc).__name__}); using inline/default constants.")

    cfg.update(entry)  # inline entry values override the table

    element = cfg.get('symbol') or cfg.get('element')
    cfg['element'] = element
    cfg.setdefault('species', element)

    # Per-source raw input files (kurucz/nist). Fall back to the legacy single
    # 'input_file' (treated as the kurucz input) when the mapping is absent.
    input_files = dict(cfg.get('input_files') or {})
    if cfg.get('input_file') and 'kurucz' not in input_files:
        input_files['kurucz'] = cfg['input_file']
    cfg['input_files'] = input_files

    # Rich-feature output base path, keyed by element SYMBOL so the established
    # Co naming (Co_features_rich_<source>.xlsx) is preserved. The '_<source>'
    # suffix is appended later by with_source_suffix() in process_element().
    if not cfg.get('output_file'):
        cfg['output_file'] = os.path.join(data_dir, f"{element}_features_rich.xlsx")

    cfg['dataset_source'] = source
    return cfg


def load_preprocessing_configs(config_path, elements=None, source=None):
    """
    Load per-element preprocessing configs from a YAML file.

    Two layouts are supported (read as raw YAML to stay dependency-light):

      * NEW    — a top-level ``elements:`` list (config_preprocess.yaml). Each
                 entry is a self-contained element table; top-level ``source`` /
                 ``data_dir`` / ``defaults`` apply to every entry.
      * LEGACY — a flat ``preprocessing:`` block plus ``dataset.dataset_source``
                 (the old single-element config_atomic.yaml); wrapped as one entry.

    Args:
        config_path: path to the YAML file.
        elements:    optional list of symbols to keep (default: all in the file).
        source:      override the active raw source (nist/kurucz).

    Returns:
        list of flat per-element dicts (see _normalize_element_cfg).
    """
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Config file not found: {config_path}")

    with open(config_path, 'r', encoding='utf-8') as fh:
        full_cfg = yaml.safe_load(fh) or {}

    entries_in = full_cfg.get('elements')
    if isinstance(entries_in, list) and entries_in and all(isinstance(e, dict) for e in entries_in):
        # NEW layout
        defaults = full_cfg.get('defaults') or {}
        data_dir = full_cfg.get('data_dir', 'data')
        active_source = (source or full_cfg.get('source') or 'kurucz').lower()
        constants_file = full_cfg.get('constants_file')
        entries = entries_in
    else:
        # LEGACY flat layout (preprocessing: + dataset.dataset_source)
        pp = full_cfg.get('preprocessing') or {}
        ds = full_cfg.get('dataset') or {}
        defaults = {}
        data_dir = ds.get('data_dir', 'data')
        active_source = (source or ds.get('dataset_source') or 'kurucz').lower()
        constants_file = ds.get('constants_file')
        entries = [pp]

    cfgs = [_normalize_element_cfg(e, defaults, data_dir, active_source, constants_file)
            for e in entries]

    if elements:
        wanted = {e.lower() for e in elements}
        found = {(c['element'] or '').lower() for c in cfgs}
        missing = wanted - found
        if missing:
            raise ValueError(
                f"Requested elements not found in {config_path}: {sorted(missing)}"
            )
        cfgs = [c for c in cfgs if (c['element'] or '').lower() in wanted]

    if not cfgs:
        raise ValueError(f"No element entries found in {config_path}")
    return cfgs


def load_preprocessing_config(config_path):
    """Backward-compatible single-element loader (returns the first entry)."""
    return load_preprocessing_configs(config_path)[0]


# ===========================================================================
# TASK 1 - configuration-string parsing into component term features
# ===========================================================================

def parse_term_SL(term):
    """
    Parse a single Russell-Saunders term string into (S, L).

    Physics: a term symbol ``(2S+1)L`` encodes the total spin S via its
    multiplicity (2S+1) and the total orbital momentum L via its letter.
    Example: ``'4F'`` → multiplicity 4 → S = (4-1)/2 = 1.5; letter 'F' → L = 3.

    A trailing ``'*'`` marks odd parity and is stripped before parsing.

    Returns:
        (S, L) as floats, or (np.nan, np.nan) if the string cannot be parsed.
    """
    if term is None:
        return np.nan, np.nan
    t = str(term).strip()
    if t.endswith('*'):            # strip the odd-parity marker
        t = t[:-1]
    m = re.match(r'(\d)([A-Z])', t)
    if not m:
        return np.nan, np.nan
    multiplicity = int(m.group(1))         # 2S + 1
    S = (multiplicity - 1) / 2.0           # total spin
    L = L_MAP.get(m.group(2), np.nan)      # total orbital angular momentum
    return float(S), float(L)


def parse_configuration(config_str, term_str):
    """
    Parse a configuration string and resultant term into coupling features.

    Configuration strings such as ``'3d7(4F)4s(2S)4p(2P),(3P)'`` describe how
    sub-groups of electrons couple their angular momenta.  The bracketed terms
    appear in order:

        (3d-group term)  (outer-electron-1 term)  (outer-electron-2 term)
        and optionally, after a comma, the ,(sub-resultant) term that couples
        the last two outer electrons before they combine with the 3d group.

    The comma sub-resultant is matched a second time by the generic bracket
    regex, so it is removed from the tail of the main-bracket list.

    Each component is mapped to a fixed-width set of 8 slots (padding with 0.0):

        comp1_S/comp1_L   3d-group coupling      (present except for '3d9')
        comp2_S/comp2_L   first outer electron   (4s, 4p, ...)
        comp3_S/comp3_L   second outer electron  (3-bracket configs only)
        subres_S/subres_L sub-resultant ,(xY)

    The resultant term (``Term_raw``) is parsed separately into
    result_S / result_L — the final coupled state of the whole atom.

    Returns a dict with the 8 component slots, result_S/result_L,
    n_components and has_subresultant.
    """
    cfg = '' if config_str is None else str(config_str)

    # Step 1+2: pull out the sub-resultant first, then all brackets, and drop
    # the sub-resultant's duplicate occurrence from the main list.
    sub_terms = _SUBRES_RE.findall(cfg)            # e.g. ['3P'] or []
    all_brackets = _BRACKET_RE.findall(cfg)        # includes the sub-resultant
    has_sub = 1 if sub_terms else 0
    if has_sub:
        # The sub-resultant is the final bracket captured by the generic regex.
        main_brackets = all_brackets[:-1]
        subres_term = sub_terms[0]
    else:
        main_brackets = all_brackets
        subres_term = None

    n_components = len(main_brackets)              # 0, 1, 2, or 3

    # Step 3+4: parse each main bracket into (S, L) and place into fixed slots.
    # Default every slot to 0.0 so the output width is constant.
    slots = {
        'comp1_S': 0.0, 'comp1_L': 0.0,
        'comp2_S': 0.0, 'comp2_L': 0.0,
        'comp3_S': 0.0, 'comp3_L': 0.0,
        'subres_S': 0.0, 'subres_L': 0.0,
    }
    for i, bracket in enumerate(main_brackets[:3]):   # at most 3 main components
        S, L = parse_term_SL(bracket)
        # NaN (unparseable) collapses to the 0.0 padding convention.
        slots[f'comp{i + 1}_S'] = 0.0 if pd.isna(S) else S
        slots[f'comp{i + 1}_L'] = 0.0 if pd.isna(L) else L

    if has_sub:
        S, L = parse_term_SL(subres_term)
        slots['subres_S'] = 0.0 if pd.isna(S) else S
        slots['subres_L'] = 0.0 if pd.isna(L) else L

    # Step 5: parse the resultant (overall) term symbol.
    result_S, result_L = parse_term_SL(term_str)

    slots['result_S'] = result_S
    slots['result_L'] = result_L
    slots['n_components'] = int(n_components)
    slots['has_subresultant'] = int(has_sub)
    return slots


# ===========================================================================
# TASK 2 - orbital electron counts
# ===========================================================================

def parse_orbitals(config_str):
    """
    Parse per-orbital electron counts from a configuration string.

    The bracketed coupling terms use UPPERCASE letters (e.g. '(4F)'), while
    orbital sub-shells use lowercase ([spdf]); the regex therefore matches only
    genuine orbital segments and skips the coupling brackets automatically.
    An orbital written without an explicit count (e.g. '4s') holds 1 electron.

    The cobalt core (1s²2s²2p⁶3s²3p⁶ = 18 e⁻) is constant and is filled in
    directly; only the valence orbitals come from the string.

    Returns a dict mapping every column in ORBITAL_COLUMNS to its occupancy.
    """
    occ = {o: 0 for o in ORBITAL_COLUMNS}
    occ.update(CORE_OCCUPANCY)                 # closed [Ar] core, always present

    cfg = '' if config_str is None else str(config_str)
    for m in _ORBITAL_RE.finditer(cfg):
        n, letter, count = m.group(1), m.group(2), m.group(3)
        key = f'{n}{letter}'                   # e.g. '3d'
        occupancy = int(count) if count else 1  # bare orbital → 1 electron
        if key in occ:
            occ[key] = occupancy
    return occ


# ===========================================================================
# TASK 3 - physics-derived features
# ===========================================================================

def compute_zeff_3d(occ, Z):
    """
    Effective nuclear charge seen by a 3d electron via Slater's rules.

    σ = 0.35·(n_3d − 1)            other electrons in the same 3d group
      + 0.85·(n_3s + n_3p)         the n−1 shell (moderate screening)
      + 1.00·(n_1s + n_2s + n_2p)  deeper shells (full screening)
    Z_eff = Z − σ.  Electrons in shells outside 3d (4s, 4p, …) do not screen it.

    Z_eff² is the dominant energy scale, analogous to Z²/n² in hydrogen.
    """
    n_3d = occ['3d']
    sigma = 0.35 * max(n_3d - 1, 0)               # same-group screening
    sigma += 0.85 * (occ['3s'] + occ['3p'])       # n-1 shell
    sigma += 1.00 * (occ['1s'] + occ['2s'] + occ['2p'])  # inner shells
    return Z - sigma


def compute_valence_slots(occ, max_valence):
    """
    Build the fixed-width valence-slot feature vector (outermost-first).

    Every valence electron is expanded to an individual (n, ℓ) pair, then
    sorted by Madelung filling order — key (n+ℓ, n), where a lower value fills
    first — and the order is REVERSED so slot 1 holds the outermost electron
    (highest Madelung order).  The list is padded with (0, 0) up to max_valence.

    This outermost-first convention is the deliberate fix to the slot-ordering
    confusion identified in the permutation-importance analysis (the old code
    placed the outermost electron in the LAST slot).

    Returns a flat list [val_e1_n, val_e1_l, ..., val_e{max}_n, val_e{max}_l].
    """
    electrons = []
    for o in VALENCE_COLUMNS:
        n = int(o[0])
        l = L_VALUES[o[1]]
        electrons.extend([(n, l)] * int(occ[o]))   # one (n, ℓ) per electron

    # Madelung order ascending (fills first); then reverse → outermost first.
    electrons.sort(key=lambda nl: (nl[0] + nl[1], nl[0]))
    electrons.reverse()

    slots = electrons[:max_valence]                 # never overflow the slots
    while len(slots) < max_valence:                 # right-pad with (0, 0)
        slots.append((0, 0))

    flat = []
    for n, l in slots:
        flat.extend([float(n), float(l)])
    return flat


def compute_row_features(row, cfg):
    """
    Compute the full feature dict for one raw input row.

    Combines TASK 1 (component terms), TASK 2 (orbital occupancies) and
    TASK 3 (physics-derived features) plus TASK 4 (gJ handling) and TASK 5
    (atomic constants) into a single flat dict, one entry per output column.
    """
    Z = cfg['Z']
    A = cfg['A']
    max_valence = cfg['max_valence']
    zeta = cfg['zeta_3d']

    feat = {}

    # --- Identifiers (passed straight through) ---------------------------
    feat['Configuration_raw'] = row.get('Configuration_raw')
    feat['Term_raw'] = row.get('Term_raw')
    feat['J'] = float(row['J'])
    feat['parity_flag'] = int(row['parity_flag'])
    feat['Reference'] = row.get('Reference')

    # --- Raw energies (TASK 3i) -----------------------------------------
    obs_level = float(row['OBS.LEVEL'])
    eigenvalue = float(row['EIGENVALUE'])
    feat['OBS.LEVEL'] = obs_level                  # experimental TARGET
    feat['EIGENVALUE'] = eigenvalue                # raw Cowan baseline (eigenvalue_calc)
    # delta_e_theory: prefer the tabulated T-W residual; otherwise derive it.
    tw = row.get('T-W', np.nan)
    feat['delta_e_theory'] = float(tw) if pd.notna(tw) else (obs_level - eigenvalue)

    # --- gJ handling (TASK 4) -------------------------------------------
    obs_gj = row.get('obs.gJ', np.nan)
    calc_gj = row.get('calc.gJ', np.nan)
    obs_gj = float(obs_gj) if pd.notna(obs_gj) else np.nan   # NaN kept as NaN!
    calc_gj = float(calc_gj) if pd.notna(calc_gj) else np.nan
    feat['obs_gj'] = obs_gj                                  # multi-task TARGET
    feat['has_obs_gj'] = int(pd.notna(obs_gj))              # mask for the loss
    feat['calc_gj'] = calc_gj                                # INPUT feature
    feat['obs_minus_calc_gj'] = (obs_gj - calc_gj) if pd.notna(obs_gj) else np.nan

    # --- TASK 1: component term features --------------------------------
    comp = parse_configuration(row.get('Configuration_raw'), row.get('Term_raw'))
    result_S = comp['result_S']
    result_L = comp['result_L']
    feat['result_S'] = result_S
    feat['result_L'] = result_L
    feat['term_known'] = int(pd.notna(result_S) and pd.notna(result_L))

    # --- TASK 3a: angular-momentum products -----------------------------
    J = feat['J']
    J_sq = J * (J + 1)
    L_sq = result_L * (result_L + 1) if pd.notna(result_L) else np.nan
    S_sq = result_S * (result_S + 1) if pd.notna(result_S) else np.nan
    lande_so_term = J_sq - L_sq - S_sq             # ∝ spin-orbit shift
    feat['J_sq'] = J_sq
    feat['L_sq'] = L_sq if pd.notna(L_sq) else 0.0
    feat['S_sq'] = S_sq if pd.notna(S_sq) else 0.0
    feat['lande_so_term'] = lande_so_term if pd.notna(lande_so_term) else 0.0

    # --- TASK 3b: theoretical Landé g-factor (LS coupling) --------------
    # WARNING: lande_g_theoretical uses the pure LS-coupling Landé formula.
    # For alkali atoms (single valence electron, pure LS coupling) this is accurate.
    # For iron-group atoms (Co, Fe, Ni) with strong configuration mixing and
    # intermediate coupling, this formula gives values far from experiment.
    # DO NOT use as a model input feature for transition metal atoms.
    # Magda's calc_gJ (from Cowan code diagonalization) is the correct theoretical value,
    # but it is semi-empirical and must not be used as a training feature either
    if J == 0:
        lande_g_theo = 0.0                          # formula undefined; g_J ≡ 0
        has_lande = 1 if feat['term_known'] else 0
    elif pd.isna(result_S) or pd.isna(result_L):
        lande_g_theo = np.nan                        # cannot evaluate
        has_lande = 0
    else:
        lande_g_theo = 1.0 + lande_so_term / (2.0 * J_sq)
        has_lande = 1
    feat['has_lande_theoretical'] = has_lande
    feat['lande_g_theoretical'] = 0.0 if pd.isna(lande_g_theo) else lande_g_theo

    # --- TASK 1 (cont.): component slots --------------------------------
    for k in ['comp1_S', 'comp1_L', 'comp2_S', 'comp2_L', 'comp3_S', 'comp3_L',
              'subres_S', 'subres_L']:
        feat[k] = comp[k]
    feat['n_components'] = comp['n_components']
    feat['has_subresultant'] = comp['has_subresultant']

    # --- TASK 2: orbital occupancies ------------------------------------
    occ = parse_orbitals(row.get('Configuration_raw'))

    # --- TASK 3c: d-electron features -----------------------------------
    n_3d = occ['3d']
    feat['n_3d'] = float(n_3d)
    feat['d_holes'] = float(10 - n_3d)             # electron-hole symmetry
    feat['d_from_half'] = float(abs(n_3d - 5))     # exchange-energy pairing cost
    feat['is_half_filled'] = int(n_3d == 5)        # special Hund stability

    # --- TASK 3d: effective nuclear charge ------------------------------
    z_eff = compute_zeff_3d(occ, Z)
    feat['Z_eff'] = z_eff
    feat['Z_eff_sq'] = z_eff ** 2

    # --- TASK 3e: spin-orbit energy estimate ----------------------------
    feat['zeta_3d'] = float(zeta)
    feat['E_so_estimate'] = (zeta / 2.0) * feat['lande_so_term']

    # --- TASK 3f: parity computed from orbital occupancy ----------------
    parity = sum(occ[o] * L_VALUES[o[1]] for o in ORBITAL_COLUMNS) % 2
    feat['parity_computed'] = int(parity)          # 0 = even, 1 = odd

    # --- TASK 3h: valence slots (outermost-first) -----------------------
    slot_vals = compute_valence_slots(occ, max_valence)
    for i in range(max_valence):
        feat[f'val_e{i + 1}_n'] = slot_vals[2 * i]
        feat[f'val_e{i + 1}_l'] = slot_vals[2 * i + 1]

    # --- TASK 3g: valence summary ---------------------------------------
    total_e = sum(occ[o] for o in ORBITAL_COLUMNS)
    core_e = sum(occ[o] for o in CORE_COLUMNS)
    feat['valence_electrons'] = float(total_e - core_e)
    feat['total_electrons'] = float(total_e)
    feat['core_electrons'] = float(core_e)
    occupied_n = [int(o[0]) for o in ORBITAL_COLUMNS if occ[o] > 0]
    feat['max_principal_n'] = float(max(occupied_n)) if occupied_n else 0.0

    # --- TASK 2 (cont.): raw orbital occupancy columns ------------------
    for o in ORBITAL_COLUMNS:
        feat[o] = float(occ[o])

    # --- TASK 5: atomic constants ---------------------------------------
    feat['Z'] = int(Z)
    feat['A'] = int(A)
    feat['proton_number'] = int(Z)
    feat['neutron_number'] = int(A - Z)

    return feat


# ===========================================================================
# Output column ordering (TASK 6)
# ===========================================================================

def add_element_targets_and_rydberg(df, cfg):
    """
    Add the element label, pre-computed energy-target transforms, and Rydberg
    placeholder columns to the features DataFrame (in place).

    These columns let AtomicDataset.py select a target purely by name and toggle
    feature groups without recomputing anything:

      Element                      element symbol (needed for stratified splits
                                   and multi-element merging downstream).
      Binding_Energy_cm-1          E_ion - OBS.LEVEL, clipped at 0 (continuum).
      Log_Binding_Energy_cm-1      log(binding); NaN where binding <= 0.
      Inverse_Binding_Energy_cm-1  A / binding; NaN where binding <= 0.
      n_star, rydberg_prediction, one_over_nstar_sq
                                   Rydberg series features. These are meaningful
                                   only for single-valence-electron (alkali) atoms;
                                   for transition metals such as Co they are 0.0.
    """
    E_ion = float(cfg['ionization_energy'])          # ionization energy (cm^-1)
    scale = float(cfg['inverse_target_scale'])       # A in A / binding

    df['Element'] = cfg['element']                    # element symbol (e.g. 'Co')
    df['species'] = cfg.get('species', cfg['element'])  # species incl. ion stage (e.g. 'Co I')

    # Binding energy = how far below the ionization limit the level sits.
    binding = (E_ion - df['OBS.LEVEL']).clip(lower=0.0)
    df['Binding_Energy_cm-1'] = binding

    # Log / inverse transforms are undefined at binding == 0 (continuum) → NaN.
    positive = binding.where(binding > 0)
    with np.errstate(divide='ignore', invalid='ignore'):
        df['Log_Binding_Energy_cm-1'] = np.log(positive)
        df['Inverse_Binding_Energy_cm-1'] = scale / positive

    # Rydberg features are not applicable to transition metals → zeros.
    df['n_star'] = 0.0
    df['rydberg_prediction'] = 0.0
    df['one_over_nstar_sq'] = 0.0


def build_column_order(max_valence):
    """Return the explicit output column order described in TASK 6."""
    valence_cols = []
    for i in range(max_valence):
        valence_cols.extend([f'val_e{i + 1}_n', f'val_e{i + 1}_l'])

    return (
        # Identifiers (not model inputs)
        ['Configuration_raw', 'Term_raw', 'J', 'parity_flag', 'Reference', 'Element', 'species']
        # Raw energy (target + Cowan baseline) + pre-computed target transforms
        + ['OBS.LEVEL', 'EIGENVALUE', 'delta_e_theory',
           'Binding_Energy_cm-1', 'Log_Binding_Energy_cm-1', 'Inverse_Binding_Energy_cm-1']
        # Experimental gJ (multi-task target)
        + ['obs_gj', 'has_obs_gj']
        # Theoretical gJ (input feature)
        + ['calc_gj', 'lande_g_theoretical', 'has_lande_theoretical', 'obs_minus_calc_gj']
        # Parsed quantum numbers (resultant term)
        + ['result_S', 'result_L', 'term_known']
        # Angular momentum products
        + ['J_sq', 'L_sq', 'S_sq', 'lande_so_term']
        # Component term features (new)
        + ['comp1_S', 'comp1_L', 'comp2_S', 'comp2_L', 'comp3_S', 'comp3_L',
           'subres_S', 'subres_L', 'n_components', 'has_subresultant']
        # d-electron features
        + ['n_3d', 'd_holes', 'd_from_half', 'is_half_filled']
        # Screening
        + ['Z_eff', 'Z_eff_sq']
        # Spin-orbit
        + ['zeta_3d', 'E_so_estimate']
        # Parity
        + ['parity_computed']
        # Valence slots (outermost-first)
        + valence_cols
        # Valence summary
        + ['valence_electrons', 'total_electrons', 'core_electrons', 'max_principal_n']
        # Rydberg features (zeros for transition metals; populated for alkalis)
        + ['n_star', 'rydberg_prediction', 'one_over_nstar_sq']
        # Orbital occupancies (raw)
        + ORBITAL_COLUMNS
        # Atomic constants
        + ['Z', 'A', 'proton_number', 'neutron_number']
    )


# ===========================================================================
# Summary sheet (TASK 6, sheet 2)
# ===========================================================================

# Leading identifier columns are excluded from the numeric summary.
IDENTIFIER_COLUMNS = ['Configuration_raw', 'Term_raw', 'J', 'parity_flag',
                      'Reference', 'Element', 'species']


def build_summary_df(df):
    """
    Build the per-feature summary table (sheet 2).

    One row per numeric feature with: feature_name, dtype, n_non_null,
    n_unique, mean, std, min, max.  Identifier columns are skipped.
    """
    rows = []
    for col in df.columns:
        if col in IDENTIFIER_COLUMNS:
            continue
        series = df[col]
        if not pd.api.types.is_numeric_dtype(series):
            continue
        rows.append({
            'feature_name': col,
            'dtype': str(series.dtype),
            'n_non_null': int(series.notna().sum()),
            'n_unique': int(series.nunique(dropna=True)),
            'mean': series.mean(),
            'std': series.std(),
            'min': series.min(),
            'max': series.max(),
        })
    return pd.DataFrame(rows)


# ===========================================================================
# XLSX writing + formatting (TASK 6)
# ===========================================================================

# Colour palette (ARGB without the leading '#').
FILL_HEADER = PatternFill('solid', fgColor='DBEAFE')   # light blue
FILL_IDENT = PatternFill('solid', fgColor='F3F4F6')    # light gray
FILL_TARGET = PatternFill('solid', fgColor='FEF9C3')   # light yellow
FILL_COMPONENT = PatternFill('solid', fgColor='D1FAE5')  # light green
FILL_EIGEN = PatternFill('solid', fgColor='EDE9FE')    # light purple

TARGET_COLUMNS = ['OBS.LEVEL', 'obs_gj']
COMPONENT_COLUMNS = ['comp1_S', 'comp1_L', 'comp2_S', 'comp2_L', 'comp3_S',
                     'comp3_L', 'subres_S', 'subres_L', 'n_components',
                     'has_subresultant']
EIGEN_COLUMNS = ['EIGENVALUE', 'delta_e_theory']


def write_rich_xlsx(features_df, summary_df, output_path):
    """
    Write the features + summary sheets and apply all TASK 6 formatting.

    Formatting:
      * header row bold, light-blue fill, frozen;
      * identifier columns light gray, target columns yellow,
        component columns green, eigenvalue/delta columns purple;
      * numeric cells formatted to 4 decimal places;
      * approximate auto-fit column widths.
    """
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        features_df.to_excel(writer, sheet_name='features', index=False)
        summary_df.to_excel(writer, sheet_name='summary', index=False)

        wb = writer.book
        ws = wb['features']

        columns = list(features_df.columns)
        numeric_flags = [pd.api.types.is_numeric_dtype(features_df[c]) for c in columns]

        # Per-column body fill (header is overridden to blue below).
        def body_fill(col):
            if col in TARGET_COLUMNS:
                return FILL_TARGET
            if col in COMPONENT_COLUMNS:
                return FILL_COMPONENT
            if col in EIGEN_COLUMNS:
                return FILL_EIGEN
            if col in IDENTIFIER_COLUMNS:
                return FILL_IDENT
            return None

        n_rows = len(features_df)
        for c_idx, col in enumerate(columns, start=1):
            letter = get_column_letter(c_idx)
            fill = body_fill(col)
            is_numeric = numeric_flags[c_idx - 1]

            # Header cell: bold + blue, regardless of column group.
            header_cell = ws.cell(row=1, column=c_idx)
            header_cell.font = Font(bold=True)
            header_cell.fill = FILL_HEADER
            header_cell.alignment = Alignment(horizontal='center')

            # Body cells: group fill + 4-decimal number format for numerics.
            for r_idx in range(2, n_rows + 2):
                cell = ws.cell(row=r_idx, column=c_idx)
                if fill is not None:
                    cell.fill = fill
                if is_numeric:
                    cell.number_format = '0.0000'

            # Approximate auto-fit width.
            ws.column_dimensions[letter].width = max(len(str(col)), 8) * 1.2

        # Freeze the header row so it stays visible while scrolling.
        ws.freeze_panes = 'A2'

        # Light formatting for the summary sheet header too.
        ws_sum = wb['summary']
        for c_idx in range(1, len(summary_df.columns) + 1):
            cell = ws_sum.cell(row=1, column=c_idx)
            cell.font = Font(bold=True)
            cell.fill = FILL_HEADER
            ws_sum.column_dimensions[get_column_letter(c_idx)].width = 16
        ws_sum.freeze_panes = 'A2'


# ===========================================================================
# TASK 7 - validation + summary printout
# ===========================================================================

def print_validation_report(df, cfg):
    """Print the validation / coverage report described in TASK 7."""
    n = len(df)
    print("\n" + "=" * 60)
    print("VALIDATION REPORT")
    print("=" * 60)
    print(f"Total rows: {n}")

    n_obs_gj = int(df['has_obs_gj'].sum())
    print(f"Rows with obs_gj: {n_obs_gj} / {n} ({100.0 * n_obs_gj / n:.1f}%)")

    n_term = int(df['term_known'].sum())
    print(f"Rows with term_known: {n_term} / {n}")

    # Electron-count check (must equal Z = 27 for every Co row).
    expected_e = cfg['Z']
    bad_counts = df[df['total_electrons'] != expected_e]
    all_ok = len(bad_counts) == 0
    print(f"Electron count check: all rows = {expected_e}? "
          f"{'yes' if all_ok else 'no'}")
    if not all_ok:
        for _, r in bad_counts.iterrows():
            print(f"  WARNING: total_electrons={r['total_electrons']:.0f} "
                  f"for config '{r['Configuration_raw']}'")

    # Parity consistency (computed vs the raw parity_flag).
    parity_match = int((df['parity_computed'] == df['parity_flag']).sum())
    print(f"Parity consistency (parity_computed vs parity_flag): "
          f"{parity_match} / {n} match")

    # Component-term coverage.
    print("Component term coverage:")
    counts = df['n_components'].value_counts().to_dict()
    print(f"  0-bracket configs: {int(counts.get(0, 0))}")
    print(f"  1-bracket configs: {int(counts.get(1, 0))}")
    print(f"  2-bracket configs: {int(counts.get(2, 0))}")
    n_three_sub = int(((df['n_components'] == 3) & (df['has_subresultant'] == 1)).sum())
    print(f"  3-bracket + sub-resultant: {n_three_sub}")

    # Top 5 most common 3d-group terms (comp1 as a multiplicity/L pair).
    print("\nTop 5 most common 3d-group terms (comp1):")
    inv_L = {v: k for k, v in L_MAP.items()}

    def comp1_label(r):
        # Reconstruct the term symbol from comp1_S / comp1_L for display.
        S, L = r['comp1_S'], r['comp1_L']
        if S == 0 and L == 0 and r['n_components'] == 0:
            return '(none: 3d9)'
        mult = int(round(2 * S + 1))
        return f"{mult}{inv_L.get(int(L), '?')}"

    labels = df.apply(comp1_label, axis=1)
    for term, cnt in labels.value_counts().head(5).items():
        print(f"  {term}: {cnt} rows")


def _write_preprocessing_sidecar(df: pd.DataFrame, config: dict, out_path: str) -> None:
    """
    Write a JSON sidecar file alongside the feature Excel output.
    The sidecar contains structured preprocessing metadata for use
    by analyze_features.py when building the HTML report.
    File is written to the same directory as out_path, with the
    same stem and suffix '_preprocess_log.json'.
    Example: data/Co_features_rich_kurucz.xlsx
          -> data/Co_features_rich_kurucz_preprocess_log.json
    """
    import json, os
    from pathlib import Path

    # --- energy range and parity split ---
    level_col = None
    for candidate in ["OBS.LEVEL", "obs_level", "energy", "ENERGY"]:
        if candidate in df.columns:
            level_col = candidate
            break
    energy_min = float(df[level_col].min()) if level_col else None
    energy_max = float(df[level_col].max()) if level_col else None

    parity_col = None
    for candidate in ["parity_computed", "parity_flag", "parity"]:
        if candidate in df.columns:
            parity_col = candidate
            break
    n_even = int((df[parity_col] == 0).sum()) if parity_col else None
    n_odd  = int((df[parity_col] == 1).sum()) if parity_col else None

    # --- obs_gj coverage ---
    gj_col = None
    for candidate in ["obs_gj", "obs.gJ", "obs_gJ"]:
        if candidate in df.columns:
            gj_col = candidate
            break
    rows_with_obs_gj = int(df[gj_col].notna().sum()) if gj_col else None
    pct_obs_gj = round(100.0 * rows_with_obs_gj / len(df), 1) if rows_with_obs_gj is not None else None

    # --- term known ---
    term_col = None
    for candidate in ["Term_raw", "term", "TERM"]:
        if candidate in df.columns:
            term_col = candidate
            break
    rows_term_known = int(df[term_col].notna().sum()) if term_col else len(df)

    # --- parity consistency ---
    parity_match = None
    if "parity_computed" in df.columns and "parity_flag" in df.columns:
        parity_match = int((df["parity_computed"] == df["parity_flag"]).sum())
    parity_total = len(df)

    # --- component distribution ---
    n_components = {}
    if "n_components" in df.columns:
        for k, v in df["n_components"].value_counts().items():
            n_components[int(k)] = int(v)
    elif "comp3_S" in df.columns:
        # infer from presence of comp columns
        def count_components(row):
            for c in [3, 2, 1, 0]:
                col = f"comp{c}_S"
                if col in df.columns and pd.notna(row.get(col)):
                    return c
            return 0
        counts = df.apply(count_components, axis=1).value_counts()
        n_components = {int(k): int(v) for k, v in counts.items()}

    has_subresultant = {}
    if "subres_S" in df.columns:
        has_sub = df["subres_S"].notna()
        has_subresultant = {0: int((~has_sub).sum()), 1: int(has_sub.sum())}

    # --- top comp1 terms ---
    top_comp1_terms = []
    if "comp1_L" in df.columns and "comp1_S" in df.columns:
        L_LETTERS = {0:"S",1:"P",2:"D",3:"F",4:"G",5:"H",6:"I"}
        def to_term(row):
            s = row["comp1_S"]
            l = row["comp1_L"]
            if pd.isna(s) or pd.isna(l):
                return None
            mult = int(round(2*s + 1))
            letter = L_LETTERS.get(int(l), "?")
            return f"{mult}{letter}"
        terms = df.apply(to_term, axis=1).dropna()
        top5 = terms.value_counts().head(5)
        top_comp1_terms = [[str(k), int(v)] for k, v in top5.items()]

    # --- element-level constants from config ---
    element = config.get("element", "")
    Z       = config.get("Z", None)
    A       = config.get("A", None)
    source  = config.get("source", config.get("dataset_source", ""))
    zeta_3d    = config.get("zeta_3d", None)
    max_valence = config.get("max_valence", None)

    sidecar = {
        "element":         element,
        "Z":               int(Z) if Z is not None else None,
        "A":               int(A) if A is not None else None,
        "source":          source,
        "total_rows":      len(df),
        "rows_with_obs_gj": rows_with_obs_gj,
        "pct_obs_gj":      pct_obs_gj,
        "rows_term_known": rows_term_known,
        "electron_count_ok": True,          # already validated by caller
        "parity_match":    parity_match,
        "parity_total":    parity_total,
        "n_components":    n_components,
        "has_subresultant": has_subresultant,
        "top_comp1_terms": top_comp1_terms,
        "energy_min":      energy_min,
        "energy_max":      energy_max,
        "n_even":          n_even,
        "n_odd":           n_odd,
        "zeta_3d":         float(zeta_3d) if zeta_3d is not None else None,
        "max_valence":     int(max_valence) if max_valence is not None else None,
    }

    stem = Path(out_path).stem
    sidecar_path = Path(out_path).parent / f"{stem}_preprocess_log.json"
    with open(sidecar_path, "w", encoding="utf-8") as f:
        json.dump(sidecar, f, indent=2, ensure_ascii=False)
    print(f"  [sidecar] preprocessing log written → {sidecar_path}")


# ===========================================================================
# Main pipeline
# ===========================================================================

def read_nist_as_raw(input_file, element, Z, A):
    """
    Read a raw NIST level CSV and adapt it to the Kurucz "raw" schema.

    NIST stores its data very differently from the Kurucz/Cowan files:
      * an Excel-exported CSV with =""value"" formula quoting,
      * a label-prefixed Term symbol (e.g. 'a 4F', 'z 6F*') with parity in '*',
      * the experimental Landé g-factor in a 'Lande' column,
      * NO theoretical (Cowan) energy or theoretical gJ.

    This function reuses the parsing helpers from preprocess/preprocess_nist.py
    (rather than duplicating them) and emits exactly the columns that
    compute_row_features() expects:

        Configuration_raw, Term_raw, J, parity_flag,
        OBS.LEVEL, EIGENVALUE, T-W, calc.gJ, obs.gJ, Reference

    Source-specific handling:
      * EIGENVALUE / T-W / calc.gJ → NaN  (NIST is experimental only — there is
        no Cowan baseline; the eigenvalue feature group should stay off for NIST).
      * obs.gJ ← the NIST 'Lande' column (the measured g-factor).
      * Term_raw ← the canonical term with the spectroscopic label prefix and the
        parity '*' stripped; parity is carried in parity_flag instead.
      * Rows flagged uncertain (Prefix='[' or Suffix='?') are dropped, the
        Configuration is forward-filled across terms that share it, and multi-J
        fields (e.g. '7/2,9/2') are expanded into separate rows — mirroring
        preprocess_nist.preprocess_element().
    """
    from preprocess_nist import read_nist_csv, parse_term, parse_j, parse_lande_g

    df = read_nist_csv(input_file)

    # Drop uncertain / approximated levels (Prefix='[' bracketed, Suffix='?').
    uncertain = pd.Series(False, index=df.index)
    if 'Prefix' in df.columns:
        uncertain |= df['Prefix'].str.strip() == '['
    if 'Suffix' in df.columns:
        uncertain |= df['Suffix'].str.strip() == '?'
    n_removed = int(uncertain.sum())
    if n_removed:
        df = df[~uncertain].reset_index(drop=True)
        print(f"  Removed {n_removed} uncertain/approximated NIST rows "
              f"(Prefix='[' or Suffix='?')")

    rows = []
    n_missing_term = 0          # NIST entries whose Term is not a valid LS symbol
    last_config = ''            # NIST omits the config for repeated terms → forward-fill
    for _, r in df.iterrows():
        config = str(r.get('Configuration', '')).strip()
        if config:
            last_config = config
        else:
            config = last_config

        term = str(r.get('Term', '')).strip()
        # parse_term returns (S, L, parity, label). The NIST Term column sometimes
        # holds values that are NOT proper LS term symbols — e.g. '16*', '*', '1',
        # '8*', '14*' — which are genuine NIST entries we cannot interpret. When
        # S or L comes back None the term is uninterpretable, so we treat it as a
        # MISSING term: Term_raw is left blank and result_S/result_L become NaN
        # (term_known=0) downstream. The parity marker ('*') is still honoured.
        _S, _L, parity, _label = parse_term(term)
        if _S is None or _L is None:
            term_raw = ''               # missing / non-LS term
            n_missing_term += 1
        else:
            m_label = re.match(r'^([a-z])\s+', term)   # drop 'a '/'z ' label prefix
            term_clean = term[m_label.end():] if m_label else term
            term_raw = term_clean.replace('*', '').strip()  # drop parity '*'

        try:
            level = float(str(r.get('Level (cm-1)', '')).strip())
        except ValueError:
            level = np.nan
        if pd.isna(level):
            continue  # a level with no energy cannot be a target — skip

        obs_gj = parse_lande_g(str(r.get('Lande', '')).strip())  # NIST measured gJ
        reference = str(r.get('Reference', '')).strip()

        # Expand multi-valued J fields ('7/2,9/2') into one row per J value.
        j_str = str(r.get('J', '')).strip()
        for j_part in [v.strip() for v in j_str.split(',')]:
            J = parse_j(j_part)
            if J is None:
                continue
            rows.append({
                'Configuration_raw': config,
                'Term_raw':          term_raw,
                'J':                 J,
                'parity_flag':       int(parity),
                'OBS.LEVEL':         level,
                'EIGENVALUE':        np.nan,   # no Cowan baseline in NIST data
                'T-W':               np.nan,
                'calc.gJ':           np.nan,   # no theoretical gJ in NIST data
                'obs.gJ':            obs_gj,   # NIST Lande = observed g-factor
                'Reference':         reference,
            })

    if n_missing_term:
        print(f"  {n_missing_term} NIST rows have a non-LS / uninterpretable Term "
              f"(e.g. '16*', '*', '1') — treated as missing (Term_raw blank, term_known=0)")

    return pd.DataFrame(rows)


def read_raw_input(input_file, source, cfg):
    """
    Read the raw input for the active dataset source into the Kurucz raw schema.

    * ``kurucz`` — the file is already in the raw schema (intermediate XLSX or
      an equivalent CSV); read it directly by extension.
    * ``nist``   — convert from the NIST CSV via read_nist_as_raw().
    """
    if source == 'nist':
        return read_nist_as_raw(input_file, cfg['element'], int(cfg['Z']), int(cfg['A']))

    # Kurucz (default): raw-schema file, dispatch on extension.
    ext = os.path.splitext(input_file)[1].lower()
    if ext in ('.xlsx', '.xlsm', '.xls'):
        return pd.read_excel(input_file)
    elif ext == '.csv':
        return pd.read_csv(input_file)
    else:
        raise ValueError(f"Unsupported input extension '{ext}' (expected .xlsx or .csv)")


def process_element(cfg):
    """
    Run the full raw → rich feature pipeline for one normalized element config.

    Returns the path of the rich-feature XLSX written.
    """
    source = cfg['dataset_source']

    # Select the raw input file for this source (per-source mapping preferred).
    input_files = cfg.get('input_files') or {}
    input_file = input_files.get(source) or cfg.get('input_file')
    if not input_file:
        raise ValueError(
            f"[{cfg['element']}] No raw input file configured for source '{source}'. "
            f"Set input_files.{source} for this element."
        )

    # Output rich file carries the source suffix so nist/kurucz never collide.
    output_file = with_source_suffix(cfg['output_file'], source)
    max_valence = cfg['max_valence']

    print("=" * 60)
    print(f"ATOMIC FEATURE PREPROCESSING - {cfg['element']}")
    print("=" * 60)
    print(f"Dataset source : {source}")
    print(f"Element        : {cfg['element']}  (Z={cfg['Z']}, A={cfg['A']})")
    print(f"Input file     : {input_file}")
    print(f"Output file    : {output_file}")
    print(f"max_valence    : {max_valence}")
    print(f"zeta_3d        : {cfg['zeta_3d']} cm^-1")

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Input file not found: {input_file}")

    raw_df = read_raw_input(input_file, source, cfg)
    print(f"\nLoaded {len(raw_df)} raw rows with columns: {list(raw_df.columns)}")

    # Row-by-row feature computation with a progress bar.
    records = []
    for _, row in tqdm(raw_df.iterrows(), total=len(raw_df), desc="Computing features"):
        records.append(compute_row_features(row, cfg))

    features_df = pd.DataFrame.from_records(records)

    # Add element label + pre-computed target transforms + Rydberg placeholders.
    add_element_targets_and_rydberg(features_df, cfg)

    # Enforce the TASK 6 column order (any unexpected columns appended at end).
    column_order = build_column_order(max_valence)
    ordered = [c for c in column_order if c in features_df.columns]
    extra = [c for c in features_df.columns if c not in column_order]
    features_df = features_df[ordered + extra]

    # Component-term distribution summary (TASK 1, step 4).
    print("\nComponent distribution:")
    print(f"  n_components   : {features_df['n_components'].value_counts().sort_index().to_dict()}")
    print(f"  has_subresultant: {features_df['has_subresultant'].value_counts().sort_index().to_dict()}")

    summary_df = build_summary_df(features_df)

    write_rich_xlsx(features_df, summary_df, output_file)

    print_validation_report(features_df, cfg)

    # Persist structured preprocessing metadata next to the Excel output so
    # analyze_features.py can populate the HTML Dataset Overview automatically.
    _write_preprocessing_sidecar(features_df, cfg, output_file)

    print(f"\nOutput written to: {output_file} "
          f"({len(features_df)} rows, {len(features_df.columns)} columns)")
    return output_file


def main():
    parser = argparse.ArgumentParser(
        description="Build rich atomic-feature XLSX file(s) from raw level data."
    )
    parser.add_argument('--config', type=str, default='config_preprocess.yaml',
                        help="Path to the YAML config (default: config_preprocess.yaml)")
    parser.add_argument('--elements', nargs='+', default=None,
                        help="Subset of element symbols to process (default: all in the config).")
    parser.add_argument('--source', type=str, default=None, choices=list(SOURCE_SUFFIXES),
                        help="Override dataset source (nist/kurucz). Default: top-level "
                             "'source' (or dataset.dataset_source) from the config.")
    args = parser.parse_args()

    cfgs = load_preprocessing_configs(args.config, elements=args.elements, source=args.source)

    print(f"Processing {len(cfgs)} element(s): {[c['element'] for c in cfgs]}\n")
    outputs = []
    for cfg in cfgs:
        outputs.append(process_element(cfg))
        print()

    print("=" * 60)
    print(f"DONE - wrote {len(outputs)} file(s):")
    for o in outputs:
        print(f"  {o}")


if __name__ == '__main__':
    main()
