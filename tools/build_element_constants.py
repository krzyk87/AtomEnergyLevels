"""
build_element_constants.py  (one-time bootstrap)

Generate data/element_constants.xlsx — the single master table of per-species
physical constants — by consolidating the values that were previously hardcoded
in two places:

  * config_preprocess.yaml      (Co, Li, Na, K, Rb, Cs, Fr: Z/A/E_ion/max_valence/zeta_3d)
  * AtomicDataset.IONIZATION_ENERGIES  (adds Fe, Ni; and the Co value 63564.6)

Where the two disagreed (Co ionization energy: 63564.0 vs 63564.6) the NIST value
63564.6 is used. Only neutral species (stage I) are seeded; add ions (e.g. 'Co II')
as new rows over time.

Run once::

    python tools/build_element_constants.py
"""

import os
import sys

import pandas as pd

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _HAVE_OPENPYXL = True
except ImportError:
    _HAVE_OPENPYXL = False

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT = os.path.join(_ROOT, "data", "element_constants.xlsx")
SHEET = "species"

# species, symbol, ion_stage, Z, A, E_ion(cm^-1), max_valence, zeta_3d, is_alkali, source
SEED = [
    # --- Transition metals (zeta_3d meaningful; not alkali) ----------------
    ("Co I", "Co", "I", 27, 59, 63564.6, 9, 515.0, False, "NIST ASD"),
    ("Fe I", "Fe", "I", 26, 56, 63737.70, 8, 0.0,  False, "NIST ASD; zeta_3d TBD"),
    ("Ni I", "Ni", "I", 28, 58, 61619.77, 10, 0.0, False, "NIST ASD; zeta_3d TBD"),
    # --- Alkali atoms (single valence electron) ----------------------------
    ("Li I", "Li", "I", 3,  7,   43487.114, 1, 0.0, True, "NIST ASD"),
    ("Na I", "Na", "I", 11, 23,  41449.451, 1, 0.0, True, "NIST ASD"),
    ("K I",  "K",  "I", 19, 39,  35009.814, 1, 0.0, True, "NIST ASD"),
    ("Rb I", "Rb", "I", 37, 85,  33690.81,  1, 0.0, True, "NIST ASD"),
    ("Cs I", "Cs", "I", 55, 133, 31406.467, 1, 0.0, True, "NIST ASD"),
    ("Fr I", "Fr", "I", 87, 223, 32848.872, 1, 0.0, True, "NIST ASD"),
]

COLUMNS = ["species", "symbol", "ion_stage", "Z", "A",
           "ionization_energy_cm-1", "max_valence", "zeta_3d", "is_alkali", "source"]


def main():
    df = pd.DataFrame(SEED, columns=COLUMNS)
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name=SHEET, index=False)
        if _HAVE_OPENPYXL:
            ws = xl.sheets[SHEET]
            header_fill = PatternFill("solid", fgColor="DBEAFE")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            widths = [10, 8, 10, 6, 6, 24, 13, 10, 11, 24]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

    print(f"Wrote {len(df)} species to {OUTPUT}")
    print(df.to_string(index=False))


if __name__ == "__main__":
    sys.exit(main())
